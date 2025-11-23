"""Excel file preprocessing module to convert complex tables to 2D structures.

This module:
1. Reads original Excel files from `sheets/` without modifying them
2. Uses LLM to reconstruct complex structures into clean 2D tables
3. Writes reconstructed tables to temporary directory (git-ignored)
"""

import hashlib
import json
import logging
import os
import uuid
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Process complex Excel files and convert them to clean 2D tables.
    
    Original files are never modified. Reconstructed tables are written to
    a temporary directory for analysis.
    """

    def __init__(self, openai_client=None, temp_dir: str = "reconstructed_tables"):
        """
        Initialize the Excel processor.
        
        Args:
            openai_client: OpenAI client for LLM-based reconstruction
            temp_dir: Directory for reconstructed tables (git-ignored)
        """
        self.openai_client = openai_client
        self.temp_dir = Path(temp_dir)
        self.temp_dir.mkdir(exist_ok=True)
        logger.info(f"Reconstructed tables directory: {self.temp_dir.absolute()}")

    def process_excel_file(self, file_path: str) -> str:
        """
        Process an Excel file and create a reconstructed 2D table.
        
        The original file is never modified. A reconstructed version is
        created in the temporary directory. If a reconstructed file already
        exists and the original hasn't changed, it will be reused.
        
        Args:
            file_path: Path to the original Excel file (in `sheets/` directory)
            
        Returns:
            Path to the reconstructed Excel file in temp directory
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Excel file not found: {file_path}")
            
            # Check if reconstructed file already exists
            existing_recon = self.get_reconstructed_path(file_path)
            if existing_recon:
                # Verify original file hasn't changed
                original_stat = Path(file_path).stat()
                recon_stat = Path(existing_recon).stat()
                
                # Check if original was modified after reconstruction
                if original_stat.st_mtime <= recon_stat.st_mtime:
                    logger.info(f"Using existing reconstructed file: {existing_recon}")
                    logger.info(f"  Original file unchanged since reconstruction")
                    return existing_recon
                else:
                    logger.info(f"Original file modified, regenerating reconstruction...")
                    # Remove old reconstructed file
                    Path(existing_recon).unlink()
            
            # Generate output path in temp directory (deterministic based on file path)
            original_name = Path(file_path).stem
            # Use a hash of the file path for deterministic naming
            file_hash = hashlib.md5(str(Path(file_path).absolute()).encode()).hexdigest()[:8]
            output_path = self.temp_dir / f"{original_name}_reconstructed_{file_hash}.xlsx"
            
            logger.info(f"Processing Excel file: {file_path}")
            logger.info(f"Original file will not be modified")
            logger.info(f"Reconstructed file will be saved to: {output_path}")
            
            # Step 1: Preprocessing - Resolving Merged Cells
            # Unmerge cells and fill blanks
            unmerged_file, merged_info = self._step1_unmerge_and_fill(file_path)
            
            try:
                # Step 2: Model Analysis - Identifying Headers and Invalid Information
                # Extract samples (first 10 rows) and use LLM to identify labels and headers
                analysis_result = self._step2_model_analysis(unmerged_file, merged_info)
                
                # Step 3: Automated Processing - Cleaning and Merging
                # Delete labels and merge multi-level headers
                reconstructed_data = self._step3_automated_processing(unmerged_file, analysis_result)
                
                # Write reconstructed data to temp directory
                self._write_reconstructed_file(reconstructed_data, output_path)
                
                logger.info(f"Excel processing complete. Reconstructed file: {output_path}")
                return str(output_path)
                
            finally:
                # Clean up temporary unmerged file
                if os.path.exists(unmerged_file):
                    os.remove(unmerged_file)
                    
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}", exc_info=True)
            raise

    def _step1_unmerge_and_fill(self, file_path: str) -> Tuple[str, Dict]:
        """
        Step 1: Preprocessing - Resolving Merged Cells
        
        Unmerge all merged cells and fill blank cells with values from merged cells.
        
        Args:
            file_path: Path to original Excel file
            
        Returns:
            Tuple of (unmerged_file_path, merged_info_dict)
        """
        try:
            logger.info("Step 1: Unmerging cells and filling blanks...")
            
            # Create temporary file for unmerged version
            unmerged_file = self.temp_dir / f"unmerged_{uuid.uuid4().hex[:8]}.xlsx"
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            merged_info = {}
            
            for ws in wb.worksheets:
                logger.info(f"  Processing sheet: {ws.title}")
                sheet_merged_info = []
                
                # Collect merge information before unmerging
                for merged_range in list(ws.merged_cells.ranges):
                    min_row, min_col, max_row, max_col = (
                        merged_range.min_row, merged_range.min_col,
                        merged_range.max_row, merged_range.max_col
                    )
                    value = ws.cell(row=min_row, column=min_col).value
                    
                    # Store merge info (for LLM analysis)
                    sheet_merged_info.append({
                        "range": str(merged_range),
                        "start": (min_row, min_col),
                        "end": (max_row, max_col),
                        "value": value
                    })
                    
                    # Unmerge cells
                    ws.unmerge_cells(start_row=min_row, start_column=min_col,
                                   end_row=max_row, end_column=max_col)
                    
                    # Fill all cells in the range with the value
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            ws.cell(row=row, column=col, value=value)
                
                merged_info[ws.title] = sheet_merged_info
                logger.info(f"    Unmerged {len(sheet_merged_info)} merged cell ranges")
            
            # Save unmerged file
            wb.save(unmerged_file)
            wb.close()
            logger.info(f"Step 1 complete. Unmerged file: {unmerged_file.name}")
            
            return str(unmerged_file), merged_info
            
        except Exception as e:
            logger.error(f"Error in Step 1 (unmerge and fill): {e}", exc_info=True)
            raise

    def _step2_model_analysis(self, unmerged_file: str, merged_info: Dict) -> List[Dict]:
        """
        Step 2: Model Analysis - Identifying Headers and Invalid Information
        
        Extract first 10 rows as samples and use LLM to identify:
        - Labels: Rows to remove (titles, notes, descriptions)
        - Multi-level Headers: Rows that form complex headers
        
        Args:
            unmerged_file: Path to unmerged Excel file
            merged_info: Dictionary with merged cell information
            
        Returns:
            List of analysis results per sheet
        """
        if self.openai_client is None:
            logger.warning("No OpenAI client provided, using default analysis")
            # Default: assume first row is header, no labels
            all_sheets = pd.read_excel(unmerged_file, sheet_name=None, header=None)
            return [
                {sheet_name: {"labels": [], "header": [1]}}
                for sheet_name in all_sheets.keys()
            ]
        
        try:
            logger.info("Step 2: Model analysis - Identifying labels and headers...")
            
            # Extract first 10 rows as samples
            excel_info = self._get_excel_data(unmerged_file, head=10)
            
            # Prepare merged cell info for LLM
            merged_info_json = json.dumps(merged_info, ensure_ascii=False, indent=2)
            
            system_prompt = '''You are a professional structured data processing AI specialized in Excel table structure analysis.

Key capabilities:
1. Accurately identify header rows: Headers are typically short labels (1-5 words) that describe data columns. They appear at the top of tables and have consistent structure.
2. Distinguish headers from data: Data rows contain actual values (numbers, dates, long descriptions, product names with details). Headers are concise column labels.
3. Identify multi-level headers: Only consecutive rows at the top that are clearly header labels (not data) should be considered headers.
4. Identify label rows: Sheet-level descriptions, titles, or notes that appear before the actual data table (not data row content).

Critical rules:
- Headers are SHORT labels (typically 1-5 words per cell), not long descriptions or data values
- If a row contains long text, numbers, or detailed product information, it is DATA, not a header
- Multi-level headers are usually 1-3 rows maximum, all at the very top
- Data rows must NEVER be identified as headers'''
            
            user_prompt = f'''Please analyze the structure of each worksheet and identify:
1. Label rows: Sheet-level titles/notes/descriptions to remove (before the actual table)
2. Header rows: The actual column header row(s) - these should be SHORT labels, not data

Important: Headers are concise column labels. If a row contains long descriptions, product details, or actual data values, it is a DATA ROW, not a header.

Data to analyze:

1. Excel file data after unmerging cells (first 10 rows):
```
{excel_info}
```

2. Original merged cell information (for determining header levels):
```
{merged_info_json}
```

Output format:
[
    {{
        "sheet_name1": {{
            "labels": [row_numbers],    # Label text rows for entire worksheet (empty list if none)
            "header": [row_numbers]      # Multi-level header rows (must include at least 1 row)
        }},
        "sheet_name2": {{
            "labels": [row_numbers],
            "header": [row_numbers]
        }}
    }}
]

Critical guidelines:
1. Headers are SHORT labels (1-5 words per cell). Long text = data row, not header.
2. Headers typically appear in the first 1-3 rows. If you see actual data values (numbers, product names with details), that's a data row.
3. Multi-level headers are rare - usually only 1-2 rows. Only mark multiple rows as headers if they are clearly all header labels (short, descriptive).
4. When in doubt, use fewer header rows. It's better to have 1 header row than to merge data into headers.
5. Label rows are sheet-level descriptions/titles before the table, not data content.
6. Each sheet analyzed independently.
7. Sheet names must match exactly.
8. Output only the JSON result, no explanations.

Example correct output:
[
    {{"sheet_name1": {{
        "labels": [1, 2],
        "header": [3, 4, 5]
    }}}},
    {{"sheet_name2": {{
        "labels": [],
        "header": [1, 2]
    }}}},
    {{"sheet_name3": {{
        "labels": [],
        "header": [1]
    }}}}
]'''
            
            response = self.openai_client.chat.completions.create(
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.3
            )
            
            result_text = response.choices[0].message.content.strip()
            # Clean up JSON if wrapped in markdown code blocks
            result_text = result_text.replace('```json', '').replace('```', '').strip()
            
            analysis_result = json.loads(result_text)
            logger.info(f"Step 2 complete. Analyzed {len(analysis_result)} sheet(s)")
            
            return analysis_result
            
        except Exception as e:
            logger.error(f"Error in Step 2 (model analysis): {e}", exc_info=True)
            logger.info("Falling back to default analysis")
            # Default: assume first row is header, no labels
            all_sheets = pd.read_excel(unmerged_file, sheet_name=None, header=None)
            return [
                {sheet_name: {"labels": [], "header": [1]}}
                for sheet_name in all_sheets.keys()
            ]

    def _step3_automated_processing(self, unmerged_file: str, analysis_result: List[Dict]) -> Dict:
        """
        Step 3: Automated Processing - Cleaning and Merging
        
        Based on LLM analysis results:
        1. Delete label rows
        2. Merge multi-level headers into single-row headers
        
        Args:
            unmerged_file: Path to unmerged Excel file
            analysis_result: List of analysis results from Step 2
            
        Returns:
            Dictionary mapping sheet names to cleaned DataFrames
        """
        try:
            logger.info("Step 3: Automated processing - Cleaning and merging headers...")
            
            reconstructed_data = {}
            
            for sheet_config in analysis_result:
                for sheet_name, config in sheet_config.items():
                    labels = config.get('labels', [])
                    header = config.get('header', [1])
                    
                    logger.info(f"  Processing sheet: {sheet_name}")
                    logger.info(f"    Labels to remove: {labels}")
                    logger.info(f"    Header rows: {header}")
                    
                    # Step 3.1: Delete label rows first (before reading with header)
                    # Read without header to manipulate rows
                    df_raw = pd.read_excel(unmerged_file, sheet_name=sheet_name, header=None, dtype=object)
                    
                    if labels:
                        # Convert 1-based to 0-based and remove label rows
                        labels_0_based = [x - 1 for x in labels]
                        df_raw = df_raw.drop(labels_0_based, axis=0, errors='ignore')
                        df_raw = df_raw.reset_index(drop=True)
                        logger.info(f"    Removed {len(labels)} label row(s)")
                    
                    # Step 3.2: Adjust header row numbers after removing labels
                    header_0_based = self._adjust_header_indices(header, labels, len(df_raw))
                    
                    if header_0_based is None:
                        # Empty dataframe or invalid headers
                        reconstructed_data[sheet_name] = pd.DataFrame()
                        continue
                    
                    # Step 3.3: Set header rows and extract data
                    df = self._extract_data_with_headers(df_raw, header_0_based)
                    
                    # Clean column names and data
                    df.columns = self._clean_column_names(df.columns)
                    
                    if len(header_0_based) > 1:
                        logger.info(f"    Merged {len(header_0_based)} header rows into single row")
                    
                    # Remove completely empty rows
                    df = df.dropna(how='all')
                    
                    reconstructed_data[sheet_name] = df
                    logger.info(f"    Final: {len(df)} rows × {len(df.columns)} columns")
            
            logger.info(f"Step 3 complete. Processed {len(reconstructed_data)} sheet(s)")
            return reconstructed_data
            
        except Exception as e:
            logger.error(f"Error in Step 3 (automated processing): {e}", exc_info=True)
            raise

    def _get_excel_data(self, file_path: str, head: int = 10) -> str:
        """
        Get sample data from Excel file for LLM analysis.
        
        Args:
            file_path: Path to Excel file
            head: Number of rows to extract per sheet
            
        Returns:
            Formatted string with sheet information
        """
        try:
            all_sheets_data = pd.read_excel(file_path, sheet_name=None, header=None)
            prompt_parts = []
            
            for sheet_name, data in all_sheets_data.items():
                data.index = data.index + 1  # 1-based indexing
                excel_col_names = [get_column_letter(i + 1) for i in range(len(data.columns))]
                data.columns = excel_col_names
                
                # Replace newlines in strings
                data = data.map(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)
                
                sheet_sample = data.head(head).to_markdown(index=True)
                sheet_info = f"Sheet: {sheet_name}\nFirst {head} rows:\n\n{sheet_sample}\n\n---"
                prompt_parts.append(sheet_info)
            
            return '\n'.join(prompt_parts)
            
        except Exception as e:
            logger.error(f"Error extracting Excel data: {e}", exc_info=True)
            raise

    def _write_reconstructed_file(self, reconstructed_data: Dict, output_path: Path) -> None:
        """
        Write reconstructed data to Excel file.
        
        Args:
            reconstructed_data: Dictionary mapping sheet names to DataFrames
            output_path: Path to output file
        """
        try:
            logger.info(f"Writing reconstructed file: {output_path}")
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in reconstructed_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"  - Sheet '{sheet_name}': {len(df)} rows × {len(df.columns)} columns")
            
            logger.info(f"Reconstructed file saved successfully")
            
        except Exception as e:
            logger.error(f"Error writing reconstructed file: {e}", exc_info=True)
            raise

    def get_reconstructed_path(self, original_path: str) -> Optional[str]:
        """
        Get path to reconstructed file if it exists.
        
        Args:
            original_path: Path to original file
            
        Returns:
            Path to reconstructed file or None if not found
        """
        original_name = Path(original_path).stem
        
        # Look for existing reconstructed file
        for file in self.temp_dir.glob(f"{original_name}_reconstructed_*.xlsx"):
            return str(file)
        
        return None

    def _adjust_header_indices(self, header: List[int], labels: List[int], df_length: int) -> Optional[List[int]]:
        """
        Adjust header row indices after removing label rows.
        
        Args:
            header: Original header row numbers (1-based)
            labels: Label row numbers that were removed (1-based)
            df_length: Length of dataframe after removing labels
            
        Returns:
            Adjusted header indices (0-based) or None if invalid
        """
        if df_length == 0:
            return None
        
        # Adjust header indices for removed labels
        if labels:
            header_adjusted = [h - sum(1 for l in labels if l < h) for h in header]
        else:
            header_adjusted = header
        
        # Convert to 0-based and validate bounds
        header_0_based = [h - 1 for h in header_adjusted if h > 0]
        header_0_based = [h for h in header_0_based if 0 <= h < df_length]
        
        if not header_0_based:
            logger.warning("    No valid header rows found, using default")
            return [0] if df_length > 0 else None
        
        return header_0_based

    def _extract_data_with_headers(self, df_raw: pd.DataFrame, header_0_based: List[int]) -> pd.DataFrame:
        """
        Extract data from dataframe using specified header rows.
        
        Args:
            df_raw: Raw dataframe without headers
            header_0_based: Header row indices (0-based)
            
        Returns:
            DataFrame with headers set and data extracted
        """
        if len(header_0_based) == 1:
            # Single-level header
            header_idx = header_0_based[0]
            if header_idx >= len(df_raw):
                # Header index out of bounds, use default column names
                df = df_raw.copy()
                df.columns = [f'Column_{i}' for i in range(len(df.columns))]
                return df
            
            df_raw.columns = df_raw.iloc[header_idx]
            data_start = header_idx + 1
            
            if data_start < len(df_raw):
                return df_raw.iloc[data_start:].reset_index(drop=True)
            else:
                # No data rows after header
                return pd.DataFrame(columns=df_raw.iloc[header_idx])
        else:
            # Multi-level header: merge into single row
            header_rows_data = df_raw.iloc[header_0_based]
            new_columns = []
            
            for col_idx in range(len(df_raw.columns)):
                col_values = []
                for row_idx in range(len(header_0_based)):
                    val = header_rows_data.iloc[row_idx, col_idx]
                    if pd.notna(val) and str(val).strip() and 'Unnamed' not in str(val):
                        val_str = str(val).strip()
                        # Only include short values (likely headers, not data)
                        if len(val_str) <= 50:  # Headers are typically short
                            col_values.append(val_str)
                
                # Remove duplicates while preserving order
                col_values_dedup = list(OrderedDict.fromkeys(col_values))
                # Only join if we have reasonable header values
                if col_values_dedup and all(len(v) <= 30 for v in col_values_dedup):
                    col_name = '-'.join(col_values_dedup)
                elif col_values_dedup:
                    # Use the shortest/most likely header value
                    col_name = min(col_values_dedup, key=len)
                else:
                    col_name = f'Column_{col_idx}'
                new_columns.append(col_name)
            
            df_raw.columns = new_columns
            data_start = max(header_0_based) + 1
            return df_raw.iloc[data_start:].reset_index(drop=True)

    def _clean_column_names(self, columns: pd.Index) -> List[str]:
        """
        Clean and standardize column names.
        
        Args:
            columns: Original column index
            
        Returns:
            List of cleaned column names
        """
        return [
            str(col).strip() if col and str(col).strip() else f'Column_{i}'
            for i, col in enumerate(columns)
        ]

    def cleanup_old_reconstructed_files(self, max_age_days: int = 7) -> None:
        """
        Clean up old reconstructed files.
        
        Args:
            max_age_days: Maximum age in days before deletion
        """
        try:
            import time
            current_time = time.time()
            max_age_seconds = max_age_days * 24 * 60 * 60
            
            for file in self.temp_dir.glob("*.xlsx"):
                file_age = current_time - file.stat().st_mtime
                if file_age > max_age_seconds:
                    logger.info(f"Removing old reconstructed file: {file}")
                    file.unlink()
                    
        except Exception as e:
            logger.warning(f"Error cleaning up old files: {e}")
